import json
import os
from typing import Any, Dict, Generator, List, Optional

from dingo.config import InputArgs
from dingo.data.datasource.base import DataSource


@DataSource.register()
class LocalDataSource(DataSource):
    def __init__(
        self,
        input_args: InputArgs = None,
        config_name: Optional[str] = None,
    ):
        """Create a `LocalDataSource` instance.
        Args:
            input_args: A `InputArgs` instance to load the dataset from.
            config_name: The name of the Hugging Face dataset configuration.
        """
        self.path = input_args.input_path
        self.config_name = config_name
        super().__init__(input_args=input_args)

    @staticmethod
    def get_source_type() -> str:
        return "local"

    def to_dict(self) -> Dict[str, Any]:
        return {
            "path": self.path,
            "config_name": self.config_name,
        }

    def load(self, **kwargs) -> Generator[str, None, None]:
        """Load the local file dataset based on `LocalDataSource`.
        Args:
            kwargs: Additional keyword arguments used for loading the dataset.
        Returns:
            An instance of `Iterable`.
        """
        return self._load_local_file()

    def _find_all_files(self, path: str, file_list: List[str]):
        """
        Find all files in path recursively.
        Args:
            path (str): The path to find all files in.
            file_list (List[str]): The list of files to find.
        """
        for _f in os.listdir(path):
            f = os.path.join(path, _f)
            if os.path.isfile(f):
                file_list.append(f)
            if os.path.isdir(f):
                self._find_all_files(f, file_list)

    def _load_excel_file_xlsx(self, path: str) -> Generator[str, None, None]:
        """
        Load an .xlsx Excel file and return its contents row by row as JSON strings.
        Args:
            path (str): The path to the Excel file.

        Returns:
            Generator[str]: Each row as a JSON string with header keys.
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl is missing. Please install it using: pip install openpyxl")

        try:
            # 使用只读模式加载工作簿，节省内存
            wb = load_workbook(filename=path, read_only=True, data_only=True)

            sheet_name = self.input_args.dataset.excel_config.sheet_name
            has_header = self.input_args.dataset.excel_config.has_header

            # 选择工作表
            if isinstance(sheet_name, str):
                if sheet_name not in wb.sheetnames:
                    raise RuntimeError(f'Sheet "{sheet_name}" not found in Excel file. Available sheets: {wb.sheetnames}')
                ws = wb[sheet_name]
            elif isinstance(sheet_name, int):
                if sheet_name >= len(wb.sheetnames):
                    raise RuntimeError(f'Sheet index {sheet_name} out of range. Total sheets: {len(wb.sheetnames)}')
                ws = wb[wb.sheetnames[sheet_name]]
            else:
                raise RuntimeError(f'Invalid sheet_name type: {type(sheet_name)}. Expected str or int.')

            # 获取所有行的迭代器
            rows = ws.iter_rows(values_only=True)

            # 处理标题行
            if has_header:
                # 读取第一行作为标题
                headers = next(rows, None)
                if headers is None:
                    wb.close()
                    raise RuntimeError(f'Excel file "{path}" is empty')

                # 将标题转换为列表，处理 None 值
                headers = [str(h) if h is not None else f'Column_{i}' for i, h in enumerate(headers)]
            else:
                # 不使用标题行，第一行也是数据
                first_row = next(rows, None)
                if first_row is None:
                    wb.close()
                    raise RuntimeError(f'Excel file "{path}" is empty')

                # 使用列序号作为列名
                headers = [str(i) for i in range(len(first_row))]

                # 处理第一行数据
                if not all(cell is None for cell in first_row):
                    row_dict = {}
                    for i, (header, value) in enumerate(zip(headers, first_row)):
                        row_dict[header] = value if value is not None else ""
                    yield json.dumps(row_dict, ensure_ascii=False) + '\n'

            # 逐行读取数据并转换为 JSON
            for row in rows:
                # 跳过空行
                if all(cell is None for cell in row):
                    continue

                # 将行数据与标题组合成字典
                row_dict = {}
                for i, (header, value) in enumerate(zip(headers, row)):
                    # 处理值为 None 的情况
                    row_dict[header] = value if value is not None else ""

                # 转换为 JSON 字符串并 yield
                yield json.dumps(row_dict, ensure_ascii=False) + '\n'

        except Exception as e:
            raise RuntimeError(
                f'Failed to read .xlsx file "{path}": {str(e)}. '
                f'Please ensure the file is a valid Excel file (.xlsx).'
            )
        finally:
            if wb:
                wb.close()

    def _load_parquet_file(self, path: str) -> Generator[str, None, None]:
        """
        Load a Parquet file and return its contents row by row as JSON strings.
        Supports streaming for large files to avoid memory overflow.

        Args:
            path (str): The path to the Parquet file.

        Returns:
            Generator[str]: Each row as a JSON string with column keys.
        """
        try:
            import pyarrow.parquet as pq
        except ImportError:
            raise RuntimeError(
                "pyarrow is required to read Parquet files. "
                "Please install it using: pip install pyarrow"
            )

        # 获取 Parquet 配置
        batch_size = self.input_args.dataset.parquet_config.batch_size
        columns = self.input_args.dataset.parquet_config.columns

        try:
            # 打开 Parquet 文件
            parquet_file = pq.ParquetFile(path)

            # 使用流式读取，分批次处理
            for batch in parquet_file.iter_batches(batch_size=batch_size, columns=columns):
                # 将 batch 转换为字典格式
                batch_dict = batch.to_pydict()

                # 获取批次中的行数
                num_rows = len(next(iter(batch_dict.values()))) if batch_dict else 0

                # 逐行处理
                for i in range(num_rows):
                    # 构建每一行的字典
                    row_dict = {col: batch_dict[col][i] for col in batch_dict}

                    # 处理特殊类型的值
                    for key, value in row_dict.items():
                        # 处理 None 值
                        if value is None:
                            row_dict[key] = ""
                        # 处理 bytes 类型
                        elif isinstance(value, bytes):
                            try:
                                row_dict[key] = value.decode('utf-8')
                            except UnicodeDecodeError:
                                row_dict[key] = str(value)
                        # 处理其他不可 JSON 序列化的类型
                        elif not isinstance(value, (str, int, float, bool, list, dict)):
                            row_dict[key] = str(value)

                    # 转换为 JSON 字符串并 yield
                    yield json.dumps(row_dict, ensure_ascii=False) + '\n'

        except ImportError as ie:
            raise RuntimeError(
                f'Failed to load required library for Parquet: {str(ie)}. '
                f'Please install pyarrow using: pip install pyarrow'
            )
        except Exception as e:
            raise RuntimeError(
                f'Failed to read Parquet file "{path}": {str(e)}. '
                f'Please ensure the file is a valid Parquet file.'
            )

    def _load_csv_file(self, path: str) -> Generator[str, None, None]:
        """
        Load a CSV file and return its contents row by row as JSON strings.
        Supports streaming for large files, different encodings, and various CSV formats.

        Args:
            path (str): The path to the CSV file.

        Returns:
            Generator[str]: Each row as a JSON string with header keys.
        """
        import csv

        # 获取 CSV 配置
        has_header = self.input_args.dataset.csv_config.has_header
        encoding = self.input_args.dataset.csv_config.encoding
        dialect = self.input_args.dataset.csv_config.dialect
        delimiter = self.input_args.dataset.csv_config.delimiter
        quotechar = self.input_args.dataset.csv_config.quotechar

        try:
            # 尝试使用指定的编码打开文件
            with open(path, 'r', encoding=encoding, newline='') as csvfile:
                # 设置 CSV reader 参数
                reader_kwargs = {
                    'dialect': dialect,
                    'quotechar': quotechar,
                }

                # 如果指定了自定义分隔符，覆盖 dialect 的默认值
                if delimiter is not None:
                    reader_kwargs['delimiter'] = delimiter

                # 创建 CSV reader（流式读取）
                csv_reader = csv.reader(csvfile, **reader_kwargs)

                # 处理标题行
                headers = None
                # first_row_data = None

                try:
                    first_row = next(csv_reader)
                except StopIteration:
                    raise RuntimeError(f'CSV file "{path}" is empty')

                if has_header:
                    # The first row is the header
                    headers = [str(h).strip() if h else f'column_{i}' for i, h in enumerate(first_row)]
                    data_rows = csv_reader
                else:
                    # Generate headers and treat the first row as data
                    from itertools import chain
                    headers = [f'column_{i}' for i in range(len(first_row))]
                    data_rows = chain([first_row], csv_reader)

                # Process all data rows in a single loop
                for row in data_rows:
                    # Skip empty rows
                    if not row or all(not cell.strip() for cell in row):
                        continue

                    # Combine row data with headers into a dictionary, handling rows with fewer columns
                    row_dict = {
                        header: (row[i].strip() if row[i] else "") if i < len(row) else ""
                        for i, header in enumerate(headers)
                    }

                    # Yield the JSON string
                    yield json.dumps(row_dict, ensure_ascii=False) + '\n'

        except UnicodeDecodeError as e:
            # 编码错误提示
            raise RuntimeError(
                f'Failed to read CSV file "{path}" with encoding "{encoding}": {str(e)}. '
                f'Please try a different encoding (e.g., "gbk", "gb2312", "latin1", "iso-8859-1").'
            )
        except csv.Error as e:
            # CSV 格式错误
            raise RuntimeError(
                f'Failed to parse CSV file "{path}": {str(e)}. '
                f'Current dialect: "{dialect}". You may need to adjust the dialect or delimiter parameter.'
            )
        except Exception as e:
            raise RuntimeError(
                f'Failed to read CSV file "{path}": {str(e)}. '
                f'Please ensure the file is a valid CSV file.'
            )

    def _load_excel_file_xls(self, path: str) -> Generator[str, None, None]:
        """
        Load an .xls Excel file and return its contents row by row as JSON strings.
        Args:
            path (str): The path to the Excel file.

        Returns:
            Generator[str]: Each row as a JSON string with header keys.
        """
        try:
            import xlrd
        except ImportError:
            raise RuntimeError(
                "xlrd is required to read .xls files. "
                "Please install it using: pip install xlrd"
            )

        try:
            # 打开工作簿
            wb = xlrd.open_workbook(path, on_demand=True)

            sheet_name = self.input_args.dataset.excel_config.sheet_name
            has_header = self.input_args.dataset.excel_config.has_header

            # 选择工作表
            if isinstance(sheet_name, str):
                try:
                    ws = wb.sheet_by_name(sheet_name)
                except xlrd.XLRDError:
                    raise RuntimeError(f'Sheet "{sheet_name}" not found in Excel file. Available sheets: {wb.sheet_names()}')
            elif isinstance(sheet_name, int):
                if sheet_name >= wb.nsheets:
                    raise RuntimeError(f'Sheet index {sheet_name} out of range. Total sheets: {wb.nsheets}')
                ws = wb.sheet_by_index(sheet_name)
            else:
                raise RuntimeError(f'Invalid sheet_name type: {type(sheet_name)}. Expected str or int.')

            if ws.nrows == 0:
                raise RuntimeError(f'Excel file "{path}" is empty')

            # 处理标题行
            start_row = 0
            if has_header:
                # 读取第一行作为标题
                headers = [str(cell.value) if cell.value is not None else f'Column_{i}'
                          for i, cell in enumerate(ws.row(0))]
                start_row = 1
            else:
                # 使用列序号作为列名
                headers = [str(i) for i in range(ws.ncols)]
                start_row = 0

            # 逐行读取数据并转换为 JSON
            for row_idx in range(start_row, ws.nrows):
                row = ws.row(row_idx)

                # 跳过空行
                if all(cell.value is None or cell.value == '' for cell in row):
                    continue

                # 将行数据与标题组合成字典
                row_dict = {}
                for i, (header, cell) in enumerate(zip(headers, row)):
                    # 处理值为 None 或空的情况
                    row_dict[header] = cell.value if cell.value is not None else ""

                # 转换为 JSON 字符串并 yield
                yield json.dumps(row_dict, ensure_ascii=False) + '\n'

        except Exception as e:
            raise RuntimeError(
                f'Failed to read .xls file "{path}": {str(e)}. '
                f'Please ensure the file is a valid Excel file (.xls).'
            )
        finally:
            if wb:
                wb.release_resources()

    def _load_local_file(self) -> Generator[str, None, None]:
        """
        Load a local file and return its contents.

        Returns:
            Generator[str]: The contents of the file.
        """
        import gzip

        if not os.path.exists(self.path):
            raise RuntimeError(f'"{self.path}" is not a valid path')

        f_list = []
        if os.path.exists(self.path) and os.path.isfile(self.path):
            f_list = [self.path]
        elif os.path.exists(self.path) and os.path.isdir(self.path):
            self._find_all_files(self.path, f_list)

        by_line = self.input_args.dataset.format not in ["json", "listjson"]

        for f in f_list:
            # Check if file is CSV
            if f.endswith('.csv'):
                if self.input_args.dataset.format != 'csv':
                    raise RuntimeError(f'CSV file "{f}" is not supported. Please set dataset.format to "csv" to read CSV files.')
                yield from self._load_csv_file(f)
            # Check if file is Parquet
            elif f.endswith('.parquet'):
                if self.input_args.dataset.format != 'parquet':
                    raise RuntimeError(f'Parquet file "{f}" is not supported. Please set dataset.format to "parquet" to read Parquet files.')
                yield from self._load_parquet_file(f)
            # Check if file is Excel
            elif f.endswith('.xlsx'):
                if self.input_args.dataset.format != 'excel':
                    raise RuntimeError(f'Excel file "{f}" is not supported. Please set dataset.format to "excel" to read Excel files.')
                yield from self._load_excel_file_xlsx(f)
            elif f.endswith('.xls'):
                if self.input_args.dataset.format != 'excel':
                    raise RuntimeError(f'Excel file "{f}" is not supported. Please set dataset.format to "excel" to read Excel files.')
                yield from self._load_excel_file_xls(f)
            # Check if file is gzipped
            elif f.endswith('.gz'):
                try:
                    with gzip.open(f, 'rt', encoding='utf-8') as _f:
                        if by_line:
                            # 使用流式读取，不使用 readlines()
                            for line in _f:
                                yield line
                        else:
                            yield _f.read()
                except Exception as gz_error:
                    raise RuntimeError(
                        f'Failed to read gzipped file "{f}": {str(gz_error)}. '
                        f'Please ensure the file is a valid gzip-compressed text file.'
                    )
            else:
                # For regular files, try UTF-8 encoding
                try:
                    with open(f, "r", encoding="utf-8") as _f:
                        if by_line:
                            # 使用流式读取，不使用 readlines()
                            for line in _f:
                                yield line
                        else:
                            yield _f.read()
                except UnicodeDecodeError as decode_error:
                    raise RuntimeError(
                        f'Failed to read file "{f}": Unsupported file format or encoding. '
                        f'Dingo only supports UTF-8 text files (.jsonl, .json, .txt), CSV files (.csv), Excel files (.xlsx, .xls) and .gz compressed text files. '
                        f'Original error: {str(decode_error)}'
                    )
                except Exception as e:
                    raise RuntimeError(
                        f'Unexpected error reading file "{f}": {str(e)}. '
                        f'Please check if the file exists and is readable.'
                    )
