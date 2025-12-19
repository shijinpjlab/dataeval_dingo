"""
Excel Dataset 测试文件

测试 .xlsx 和 .xls 文件的流式读取功能
"""

import json
import os
import tempfile

from dingo.config import DatasetArgs, DatasetExcelArgs, InputArgs
from dingo.data.dataset.local import LocalDataset
from dingo.data.datasource.local import LocalDataSource


def create_test_xlsx_file(file_path: str, has_header: bool = True):
    """创建测试用的 .xlsx 文件"""
    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ openpyxl 未安装，跳过 .xlsx 文件测试")
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "测试数据"

    if has_header:
        # 添加表头
        ws.append(["姓名", "年龄", "城市", "分数"])
        # 添加数据
        ws.append(["张三", 25, "北京", 95.5])
        ws.append(["李四", 30, "上海", 88.0])
        ws.append(["王五", 28, "广州", 92.3])
        ws.append(["赵六", 35, "深圳", 87.8])
    else:
        # 直接添加数据，没有表头
        ws.append(["张三", 25, "北京", 95.5])
        ws.append(["李四", 30, "上海", 88.0])
        ws.append(["王五", 28, "广州", 92.3])
        ws.append(["赵六", 35, "深圳", 87.8])

    # 创建第二个工作表
    ws2 = wb.create_sheet("第二个表")
    ws2.append(["ID", "名称"])
    ws2.append([1, "项目A"])
    ws2.append([2, "项目B"])

    wb.save(file_path)
    wb.close()
    return True


def create_test_xls_file(file_path: str, has_header: bool = True):
    """创建测试用的 .xls 文件"""
    try:
        import xlwt
    except ImportError:
        print("⚠ xlwt 未安装，跳过 .xls 文件测试")
        return False

    wb = xlwt.Workbook()
    ws = wb.add_sheet("测试数据")

    row_idx = 0
    if has_header:
        # 添加表头
        ws.write(row_idx, 0, "姓名")
        ws.write(row_idx, 1, "年龄")
        ws.write(row_idx, 2, "城市")
        ws.write(row_idx, 3, "分数")
        row_idx += 1

    # 添加数据
    data = [
        ["张三", 25, "北京", 95.5],
        ["李四", 30, "上海", 88.0],
        ["王五", 28, "广州", 92.3],
        ["赵六", 35, "深圳", 87.8],
    ]

    for row_data in data:
        for col_idx, value in enumerate(row_data):
            ws.write(row_idx, col_idx, value)
        row_idx += 1

    # 创建第二个工作表
    ws2 = wb.add_sheet("第二个表")
    ws2.write(0, 0, "ID")
    ws2.write(0, 1, "名称")
    ws2.write(1, 0, 1)
    ws2.write(1, 1, "项目A")
    ws2.write(2, 0, 2)
    ws2.write(2, 1, "项目B")

    wb.save(file_path)
    return True


def test_xlsx_with_header():
    """测试有表头的 .xlsx 文件"""
    print("=" * 60)
    print("测试 .xlsx 文件（有表头）")
    print("=" * 60)

    # 创建临时文件
    temp_dir = tempfile.mkdtemp()
    xlsx_file = os.path.join(temp_dir, "test_data_with_header.xlsx")

    try:
        # 创建测试文件
        if not create_test_xlsx_file(xlsx_file, has_header=True):
            return

        print(f"✓ 创建测试文件: {xlsx_file}")

        # 配置参数
        excel_config = DatasetExcelArgs(
            sheet_name=0,  # 读取第一个工作表
            has_header=True  # 第一行是表头
        )

        dataset_config = DatasetArgs(
            source="local",
            format="excel",
            excel_config=excel_config
        )

        input_args = InputArgs(
            task_name="excel_test",
            input_path=xlsx_file,
            output_path="outputs/excel_test/",
            dataset=dataset_config,
            evaluator=[]
        )

        print("✓ 配置参数创建成功")

        # 创建数据源和数据集
        datasource = LocalDataSource(input_args=input_args)
        print("✓ LocalDataSource 创建成功")

        dataset = LocalDataset(source=datasource, name="test_excel_dataset")
        print("✓ LocalDataset 创建成功")

        # 流式读取数据
        print("\n开始流式读取数据:")
        count = 0
        for idx, data in enumerate(dataset.get_data()):
            count += 1
            print(f"  [{idx + 1}] {data}")

            # 验证数据格式
            if idx == 0:
                # 第一行数据应该有 "姓名", "年龄", "城市", "分数" 这些键
                assert hasattr(data, 'content'), "数据缺少 content 属性"
                data_dict = json.loads(data.content)
                assert "姓名" in data_dict, "数据缺少 '姓名' 字段"
                assert "年龄" in data_dict, "数据缺少 '年龄' 字段"
                assert "城市" in data_dict, "数据缺少 '城市' 字段"
                assert "分数" in data_dict, "数据缺少 '分数' 字段"
                print("✓ 数据格式验证通过")

        assert count == 4, f"期望读取 4 行数据，实际读取了 {count} 行"
        print(f"\n✓ 成功读取 {count} 条数据")

        print("\n" + "=" * 60)
        print("✓ 测试通过!")
        print("=" * 60)

    finally:
        # 清理临时文件
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n✓ 清理临时文件: {temp_dir}")


def test_xlsx_without_header():
    """测试无表头的 .xlsx 文件"""
    print("\n" + "=" * 60)
    print("测试 .xlsx 文件（无表头）")
    print("=" * 60)

    # 创建临时文件
    temp_dir = tempfile.mkdtemp()
    xlsx_file = os.path.join(temp_dir, "test_data_without_header.xlsx")

    try:
        # 创建测试文件
        if not create_test_xlsx_file(xlsx_file, has_header=False):
            return

        print(f"✓ 创建测试文件: {xlsx_file}")

        # 配置参数
        excel_config = DatasetExcelArgs(
            sheet_name=0,
            has_header=False  # 第一行不是表头
        )

        dataset_config = DatasetArgs(
            source="local",
            format="excel",
            excel_config=excel_config
        )

        input_args = InputArgs(
            task_name="excel_test_no_header",
            input_path=xlsx_file,
            output_path="outputs/excel_test/",
            dataset=dataset_config,
            evaluator=[]
        )

        print("✓ 配置参数创建成功")

        # 创建数据源和数据集
        datasource = LocalDataSource(input_args=input_args)
        dataset = LocalDataset(source=datasource, name="test_excel_no_header")

        # 流式读取数据
        print("\n开始流式读取数据:")
        count = 0
        for idx, data in enumerate(dataset.get_data()):
            count += 1
            print(f"  [{idx + 1}] {data}")

            # 验证数据格式（使用数字作为列名）
            if idx == 0:
                data_dict = json.loads(data.content)
                assert "0" in data_dict, "数据缺少 '0' 字段"
                assert "1" in data_dict, "数据缺少 '1' 字段"
                assert "2" in data_dict, "数据缺少 '2' 字段"
                assert "3" in data_dict, "数据缺少 '3' 字段"
                print("✓ 数据格式验证通过（使用列序号作为键）")

        assert count == 4, f"期望读取 4 行数据，实际读取了 {count} 行"
        print(f"\n✓ 成功读取 {count} 条数据")

        print("\n" + "=" * 60)
        print("✓ 测试通过!")
        print("=" * 60)

    finally:
        # 清理临时文件
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n✓ 清理临时文件: {temp_dir}")


def test_xlsx_sheet_by_name():
    """测试通过工作表名称读取"""
    print("\n" + "=" * 60)
    print("测试 .xlsx 文件（通过工作表名称读取）")
    print("=" * 60)

    # 创建临时文件
    temp_dir = tempfile.mkdtemp()
    xlsx_file = os.path.join(temp_dir, "test_data_sheet_name.xlsx")

    try:
        # 创建测试文件
        if not create_test_xlsx_file(xlsx_file, has_header=True):
            return

        print(f"✓ 创建测试文件: {xlsx_file}")

        # 配置参数 - 读取第二个工作表
        excel_config = DatasetExcelArgs(
            sheet_name="第二个表",  # 使用工作表名称
            has_header=True
        )

        dataset_config = DatasetArgs(
            source="local",
            format="excel",
            excel_config=excel_config
        )

        input_args = InputArgs(
            task_name="excel_test_sheet_name",
            input_path=xlsx_file,
            output_path="outputs/excel_test/",
            dataset=dataset_config,
            evaluator=[]
        )

        print("✓ 配置参数创建成功")

        # 创建数据源和数据集
        datasource = LocalDataSource(input_args=input_args)
        dataset = LocalDataset(source=datasource, name="test_excel_sheet_name")

        # 流式读取数据
        print("\n开始流式读取数据（第二个工作表）:")
        count = 0
        for idx, data in enumerate(dataset.get_data()):
            count += 1
            print(f"  [{idx + 1}] {data}")

            # 验证数据格式
            if idx == 0:
                data_dict = json.loads(data.content)
                assert "ID" in data_dict, "数据缺少 'ID' 字段"
                assert "名称" in data_dict, "数据缺少 '名称' 字段"
                print("✓ 数据格式验证通过")

        assert count == 2, f"期望读取 2 行数据，实际读取了 {count} 行"
        print(f"\n✓ 成功读取 {count} 条数据")

        print("\n" + "=" * 60)
        print("✓ 测试通过!")
        print("=" * 60)

    finally:
        # 清理临时文件
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n✓ 清理临时文件: {temp_dir}")


def test_xls_with_header():
    """测试有表头的 .xls 文件"""
    print("\n" + "=" * 60)
    print("测试 .xls 文件（有表头）")
    print("=" * 60)

    # 创建临时文件
    temp_dir = tempfile.mkdtemp()
    xls_file = os.path.join(temp_dir, "test_data_with_header.xls")

    try:
        # 创建测试文件
        if not create_test_xls_file(xls_file, has_header=True):
            return

        print(f"✓ 创建测试文件: {xls_file}")

        # 配置参数
        excel_config = DatasetExcelArgs(
            sheet_name=0,
            has_header=True
        )

        dataset_config = DatasetArgs(
            source="local",
            format="excel",
            excel_config=excel_config
        )

        input_args = InputArgs(
            task_name="xls_test",
            input_path=xls_file,
            output_path="outputs/xls_test/",
            dataset=dataset_config,
            evaluator=[]
        )

        print("✓ 配置参数创建成功")

        # 创建数据源和数据集
        datasource = LocalDataSource(input_args=input_args)
        print("✓ LocalDataSource 创建成功")

        dataset = LocalDataset(source=datasource, name="test_xls_dataset")
        print("✓ LocalDataset 创建成功")

        # 流式读取数据
        print("\n开始流式读取数据:")
        count = 0
        for idx, data in enumerate(dataset.get_data()):
            count += 1
            print(f"  [{idx + 1}] {data}")

            # 验证数据格式
            if idx == 0:
                data_dict = json.loads(data.content)
                assert "姓名" in data_dict, "数据缺少 '姓名' 字段"
                assert "年龄" in data_dict, "数据缺少 '年龄' 字段"
                print("✓ 数据格式验证通过")

        assert count == 4, f"期望读取 4 行数据，实际读取了 {count} 行"
        print(f"\n✓ 成功读取 {count} 条数据")

        print("\n" + "=" * 60)
        print("✓ 测试通过!")
        print("=" * 60)

    finally:
        # 清理临时文件
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n✓ 清理临时文件: {temp_dir}")


def test_stream_large_xlsx():
    """测试大文件的流式读取特性"""
    print("\n" + "=" * 60)
    print("测试流式读取特性（大文件）")
    print("=" * 60)

    temp_dir = tempfile.mkdtemp()
    xlsx_file = os.path.join(temp_dir, "large_test.xlsx")

    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ openpyxl 未安装，跳过大文件测试")
        return

    try:
        # 创建包含较多数据的测试文件
        wb = Workbook()
        ws = wb.active

        # 添加表头
        ws.append(["ID", "名称", "数值"])

        # 添加 1000 行数据
        for i in range(1, 1001):
            ws.append([i, f"项目_{i}", i * 1.5])

        wb.save(xlsx_file)
        wb.close()

        print(f"✓ 创建包含 1000 行数据的测试文件")

        # 配置参数
        excel_config = DatasetExcelArgs(
            sheet_name=0,
            has_header=True
        )

        dataset_config = DatasetArgs(
            source="local",
            format="excel",
            excel_config=excel_config
        )

        input_args = InputArgs(
            task_name="stream_test",
            input_path=xlsx_file,
            output_path="outputs/stream_test/",
            dataset=dataset_config,
            evaluator=[]
        )

        # 创建数据源和数据集
        datasource = LocalDataSource(input_args=input_args)
        dataset = LocalDataset(source=datasource, name="stream_test_dataset")

        # 只读取前 10 条，验证流式读取
        print("开始流式读取（只读取前 10 条）:")
        count = 0
        for idx, data in enumerate(dataset.get_data()):
            if idx < 10:
                print(f"  [{idx + 1}] {data}")
            count += 1
            if idx >= 9:  # 只读取前 10 条就停止
                break

        print(f"\n✓ 流式读取验证通过（处理了 {count} 条数据后停止）")

        print("\n" + "=" * 60)
        print("✓ 测试通过!")
        print("=" * 60)

    finally:
        # 清理临时文件
        import shutil
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print(f"\n✓ 清理临时文件: {temp_dir}")


if __name__ == "__main__":
    print("\n")
    print("╔" + "═" * 58 + "╗")
    print("║" + " " * 15 + "Excel 数据集测试套件" + " " * 21 + "║")
    print("╚" + "═" * 58 + "╝")
    print("\n")

    # 测试 .xlsx 文件
    test_xlsx_with_header()
    test_xlsx_without_header()
    test_xlsx_sheet_by_name()

    # 测试 .xls 文件
    test_xls_with_header()

    # 测试流式读取
    test_stream_large_xlsx()

    print("\n")
    print("╔" + "═" * 58 + "╗")
    print("║" + " " * 18 + "所有测试完成!" + " " * 23 + "║")
    print("╚" + "═" * 58 + "╝")
    print("\n")
